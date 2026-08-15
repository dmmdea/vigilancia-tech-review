#!/usr/bin/env python3
"""Make EVERY submitted file reviewable, whatever its format.

Design rule: the Read tool renders PDFs and images VISUALLY and reads text.
So anything a reviewer must SEE becomes a PDF or an image; anything that is
really data becomes text. Convert ONLY where Read cannot open the format at
all — never to "simplify" something Read already handles.

  .pdf                -> as-is                        (vision)
  .pptx .ppt .odp     -> PowerPoint/LibreOffice -> PDF (vision)
  .docx .doc .rtf     -> Word/LibreOffice -> PDF       (vision; KEEPS the
                         screenshots students embed as their PoC evidence —
                         text extraction would silently drop exactly the
                         evidence carrying the 50% PoC weight)
  .html .htm          -> headless Chrome/Edge -> PDF   (vision; keeps layout)
  .png .jpg .gif ...  -> passthrough, NO conversion    (vision)
  .xlsx .xls .csv     -> cell dump -> .txt             (data reads as data)
  .mp4 .mov ...       -> ffmpeg keyframes (vision) + optional transcript
  .zip                -> extracted, contents re-routed by these same rules
  .py .txt .md ...    -> passthrough                   (text)

Videos: frames go to the REVIEWER's own vision (grading is load-bearing, so
the strong model looks at real frames). Audio transcription is mechanical and
can be filled in separately (e.g. a local whisper) via --transcript.

Usage:
    python prepare_materials.py <workdir> [--only=<folder_id,...>]
                                [--matroot=<short dir>]
                                [--rasterize]
                                [--transcript=<folder_id>=<file.txt>]...

--rasterize: additionally run pdf_to_images.py on every produced PDF and
record `page_images` (+ `truncated_from` when the PDF exceeds the page cap)
on the item — REQUIRED on harnesses whose file reader cannot render PDF
pages (Codex, Antigravity); build_bundles --images then emits image
instructions. Without this flag those keys never exist and --images
silently degrades to PDF instructions the reviewer cannot follow.

Reads <workdir>/review_plan.json, writes <workdir>/materials.json.

WINDOWS LONG PATHS + MAX_PATH: source paths are read through the \\\\?\\
prefix, but OUTPUT goes under --matroot, which must be SHORT. The external
renderers used here (Chrome, ffmpeg, Word/Excel/PowerPoint COM) do NOT
reliably accept the \\\\?\\ prefix, so short output paths are the only robust
fix. Default matroot is a short dir in the system temp area, never the
(often very deep) session scratchpad.
"""
import json
import os
import shutil
import subprocess
import sys
import tempfile
import zipfile

try:
    from pypdf import PdfReader
except ImportError:
    print("ERROR: pypdf no está instalado (pip install -r requirements.txt). "
          "Sin él no se puede contar páginas y el gate de cobertura queda "
          "desactivado para TODOS los estudiantes — corrígelo antes de seguir.",
          file=sys.stderr)
    sys.exit(2)

for _s in (sys.stdout, sys.stderr):
    if _s in (sys.__stdout__, sys.__stderr__) and hasattr(_s, "reconfigure"):
        _s.reconfigure(encoding="utf-8")

IMAGE_EXT = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp", ".tif", ".tiff"}
VIDEO_EXT = {".mp4", ".mov", ".mkv", ".avi", ".webm", ".m4v"}
# .csv deliberately NOT here: openpyxl cannot open CSV — it must route to the
# TEXT_EXT passthrough (an earlier version sent every .csv to a guaranteed
# "lectura de hoja de cálculo falló" error item).
SHEET_EXT = {".xlsx", ".xlsm", ".xls"}
TEXT_EXT = {".py", ".txt", ".md", ".json", ".log", ".sql", ".js", ".ts", ".csv"}
DOC_EXT = {".docx", ".doc", ".rtf", ".odt"}
HTML_EXT = {".html", ".htm"}
SLIDE_EXT = {".pptx", ".ppt", ".odp"}

HERE = os.path.dirname(os.path.abspath(__file__))


def longpath(p: str) -> str:
    if os.name != "nt":
        return p
    ap = os.path.abspath(p)
    if ap.startswith("\\\\?\\"):
        return ap
    if ap.startswith("\\\\"):
        return "\\\\?\\UNC\\" + ap[2:]
    return "\\\\?\\" + ap


def find_exe(*cands):
    for c in cands:
        if os.path.sep in c or (os.name == "nt" and ":" in c):
            if os.path.exists(c):
                return c
        else:
            w = shutil.which(c)
            if w:
                return w
    return None


CHROME = find_exe(
    "chrome", "msedge",
    r"C:\Program Files\Google\Chrome\Application\chrome.exe",
    r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
    r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe")
FFMPEG = find_exe("ffmpeg")
SOFFICE = find_exe(
    "soffice", r"C:\Program Files\LibreOffice\program\soffice.exe")


def pdf_pages(path):
    """Page count, or None for THIS unreadable file (pypdf import is checked
    at module load — a missing dependency is exit 2, never a silent None)."""
    try:
        return len(PdfReader(path).pages)
    except Exception:
        return None


def stage(src, dst):
    os.makedirs(os.path.dirname(dst), exist_ok=True)
    shutil.copyfile(longpath(src), dst)
    return dst


WORD_PS = r"""
$ErrorActionPreference = 'Stop'
$w = New-Object -ComObject Word.Application
$w.Visible = $false
try {
  $d = $w.Documents.Open($env:VTR_SRC, $false, $true)
  $d.SaveAs([ref]$env:VTR_DST, [ref]17)
  $d.Close($false)
} finally { $w.Quit() }
"""


def run_ps(body, src, dst, timeout=300):
    with tempfile.NamedTemporaryFile(mode="w", suffix=".ps1", delete=False,
                                     encoding="utf-8") as f:
        f.write(body)
        script = f.name
    try:
        env = dict(os.environ, VTR_SRC=os.path.abspath(src),
                   VTR_DST=os.path.abspath(dst))
        r = subprocess.run(["powershell", "-NoProfile", "-ExecutionPolicy",
                            "Bypass", "-File", script],
                           capture_output=True, text=True, errors="replace",
                           timeout=timeout, env=env)
        if r.returncode != 0 or not os.path.exists(dst):
            return False, (r.stderr or r.stdout).strip()[:300]
        return True, ""
    except subprocess.TimeoutExpired:
        return False, "timeout"
    finally:
        os.unlink(script)


def soffice_convert(src, dst, timeout=600):
    if not SOFFICE:
        return False, "LibreOffice no disponible"
    outdir = os.path.dirname(os.path.abspath(dst))
    produced = os.path.join(outdir,
                            os.path.splitext(os.path.basename(src))[0] + ".pdf")
    try:
        r = subprocess.run([SOFFICE, "--headless", "--convert-to", "pdf",
                            "--outdir", outdir, src],
                           capture_output=True, text=True, errors="replace",
                           timeout=timeout)
    except (subprocess.TimeoutExpired, OSError) as e:
        return False, str(e)[:200]
    if not os.path.exists(produced):
        return False, (r.stderr or r.stdout).strip()[:300]
    if os.path.abspath(produced) != os.path.abspath(dst):
        shutil.move(produced, dst)
    return True, ""


def doc_to_pdf(src, dst):
    """Word COM first (best fidelity for embedded screenshots), then LO."""
    if os.name == "nt":
        ok, err = run_ps(WORD_PS, src, dst)
        if ok:
            return True, ""
    else:
        err = "no Windows"
    ok2, err2 = soffice_convert(src, dst)
    return (True, "") if ok2 else (False, f"Word: {err} | LibreOffice: {err2}")


def html_to_pdf(src, dst):
    if not CHROME:
        return soffice_convert(src, dst)
    try:
        subprocess.run(
            [CHROME, "--headless", "--disable-gpu", "--no-sandbox",
             "--run-all-compositor-stages-before-draw",
             "--virtual-time-budget=10000", "--no-pdf-header-footer",
             f"--print-to-pdf={os.path.abspath(dst)}",
             "file:///" + os.path.abspath(src).replace("\\", "/")],
            capture_output=True, text=True, errors="replace", timeout=180)
    except (subprocess.TimeoutExpired, OSError) as e:
        return False, str(e)[:200]
    return (True, "") if os.path.exists(dst) else (False, "Chrome no produjo PDF")


def slides_to_pdf(src, dst):
    r = subprocess.run([sys.executable,
                        os.path.join(HERE, "convert_to_pdf.py"), src, dst],
                       capture_output=True, text=True, errors="replace",
                       timeout=650)
    return (r.returncode == 0), r.stderr.strip()[:300]


def sheet_to_text(src, dst_txt, max_rows=400):
    """Returns a truncation note (str) or None if everything was dumped."""
    import openpyxl
    # NOT read_only: a ReadOnlyWorksheet lacks .dimensions and friends.
    wb = openpyxl.load_workbook(src, data_only=True)
    lines = []
    truncated = []
    for ws in wb.worksheets:
        lines.append(f"=== HOJA: {ws.title} "
                     f"({ws.max_row} filas x {ws.max_column} columnas) ===")
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            if i > max_rows:
                lines.append(f"... (truncado en {max_rows} de {ws.max_row} filas)")
                truncated.append(f"{ws.title}: {max_rows}/{ws.max_row} filas")
                break
            if any(c is not None and str(c).strip() for c in row):
                lines.append(" | ".join("" if c is None else str(c) for c in row))
        lines.append("")
    with open(dst_txt, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    return ("TRUNCADO — " + "; ".join(truncated)) if truncated else None


def video_keyframes(src, outdir, n=8):
    if not FFMPEG:
        return [], None, "ffmpeg no disponible"
    os.makedirs(outdir, exist_ok=True)
    probe = subprocess.run([FFMPEG, "-i", src], capture_output=True,
                           text=True, errors="replace")
    dur = None
    for line in (probe.stderr or "").splitlines():
        if "Duration:" in line:
            try:
                t = line.split("Duration:")[1].split(",")[0].strip()
                h, m, s = t.split(":")
                dur = int(h) * 3600 + int(m) * 60 + float(s)
            except Exception:
                pass
            break
    step = max(1, int((dur or 60) / n))
    r = subprocess.run([FFMPEG, "-y", "-i", src, "-vf",
                        f"fps=1/{step},scale=1280:-2", "-frames:v", str(n),
                        os.path.join(outdir, "f%02d.jpg")],
                       capture_output=True, text=True, errors="replace",
                       timeout=600)
    frames = sorted(f for f in os.listdir(outdir) if f.endswith(".jpg"))
    return frames, dur, "" if frames else (r.stderr or "")[-300:]


def rasterize_pdf(pdf_path, outdir):
    """Run pdf_to_images.py; returns (images, truncated_from|None, err)."""
    r = subprocess.run([sys.executable,
                        os.path.join(HERE, "pdf_to_images.py"),
                        pdf_path, outdir],
                       capture_output=True, text=True, errors="replace",
                       timeout=600)
    if r.returncode == 2:
        print(r.stderr, file=sys.stderr)
        sys.exit(2)          # environment problem (pymupdf missing): stop
    if r.returncode != 0:
        return [], None, r.stderr.strip()[:200]
    # tolerate library banners/warnings before the JSON line: parse the first
    # line that IS a JSON object rather than trusting stdout to be clean
    info = None
    for line in r.stdout.splitlines():
        line = line.strip()
        if line.startswith("{"):
            try:
                info = json.loads(line)
                break
            except json.JSONDecodeError:
                continue
    if info is None:
        return [], None, f"salida no-JSON de pdf_to_images: {r.stdout[:120]!r}"
    return info["images"], info.get("truncated_from"), ""


def process_file(fr, folder_id, items, matroot, prefix=""):
    name = fr["name"]
    ext = (fr.get("ext") or os.path.splitext(name)[1]).lower()
    src = fr["path"]
    base = os.path.join(matroot, folder_id)
    os.makedirs(base, exist_ok=True)
    tag = f"{prefix}{len(items):02d}"

    def ok_pdf(dst, note=None):
        items.append({"kind": "pdf", "path": os.path.abspath(dst),
                      "label": name, "pages": pdf_pages(dst),
                      **({"note": note} if note else {})})

    def fail(msg):
        items.append({"kind": "error", "label": name, "note": msg})

    try:
        if ext == ".pdf":
            ok_pdf(stage(src, os.path.join(base, f"{tag}.pdf")))
        elif ext in SLIDE_EXT:
            local = stage(src, os.path.join(base, f"{tag}_src{ext}"))
            dst = os.path.join(base, f"{tag}.pdf")
            good, err = slides_to_pdf(local, dst)
            ok_pdf(dst) if good else fail(f"conversión de diapositivas falló: {err}")
        elif ext in DOC_EXT:
            local = stage(src, os.path.join(base, f"{tag}_src{ext}"))
            dst = os.path.join(base, f"{tag}.pdf")
            good, err = doc_to_pdf(local, dst)
            (ok_pdf(dst, "documento de texto renderizado a PDF (conserva capturas)")
             if good else fail(f"conversión Word falló: {err}"))
        elif ext in HTML_EXT:
            local = stage(src, os.path.join(base, f"{tag}_src.html"))
            dst = os.path.join(base, f"{tag}.pdf")
            good, err = html_to_pdf(local, dst)
            (ok_pdf(dst, "página HTML renderizada en navegador")
             if good else fail(f"render HTML falló: {err}"))
        elif ext in IMAGE_EXT:
            dst = stage(src, os.path.join(base, f"{tag}{ext}"))
            items.append({"kind": "image", "path": os.path.abspath(dst),
                          "label": name})
        elif ext in SHEET_EXT:
            local = stage(src, os.path.join(base, f"{tag}_src{ext}"))
            dst = os.path.join(base, f"{tag}.txt")
            try:
                trunc = sheet_to_text(local, dst)
                note = "datos de hoja de cálculo extraídos"
                if trunc:
                    note += f" ({trunc}; filas restantes NO revisadas)"
                items.append({"kind": "text", "path": os.path.abspath(dst),
                              "label": name, "note": note})
            except Exception as e:
                fail(f"lectura de hoja de cálculo falló: {e}")
        elif ext in VIDEO_EXT:
            local = stage(src, os.path.join(base, f"{tag}_src{ext}"))
            fdir = os.path.join(base, f"{tag}_frames")
            frames, dur, err = video_keyframes(local, fdir)
            items.append({
                "kind": "video", "label": name,
                "video_path": os.path.abspath(local), "duration_sec": dur,
                "frames": [os.path.abspath(os.path.join(fdir, f)) for f in frames],
                "transcript_path": None,
                "note": ("fotogramas extraídos para revisión visual"
                         if frames else f"extracción de fotogramas falló: {err}")})
        elif ext in TEXT_EXT:
            dst = stage(src, os.path.join(base, f"{tag}{ext}"))
            items.append({"kind": "text", "path": os.path.abspath(dst),
                          "label": name})
        elif ext == ".zip":
            local = stage(src, os.path.join(base, f"{tag}_src.zip"))
            xdir = os.path.join(base, f"{tag}_unzipped")
            os.makedirs(xdir, exist_ok=True)
            try:
                with zipfile.ZipFile(local) as z:
                    z.extractall(xdir)
            except Exception as e:
                fail(f"no se pudo descomprimir: {e}")
                return
            inner = [os.path.join(r, fn)
                     for r, _d, fs in os.walk(xdir) for fn in sorted(fs)]
            for p in inner[:40]:
                process_file({"name": f"{name} :: {os.path.relpath(p, xdir)}",
                              "ext": os.path.splitext(p)[1].lower(), "path": p},
                             folder_id, items, matroot, prefix="z")
            if len(inner) > 40:
                items.append({"kind": "error", "label": name,
                              "note": f"zip con {len(inner)} archivos; solo se "
                                      "procesaron los primeros 40 — revisar el resto"})
        else:
            fail(f"extensión no reconocida ({ext}) — revisar manualmente")
    except Exception as e:
        fail(f"fallo procesando: {e}")


def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    if len(args) != 1:
        print(__doc__, file=sys.stderr)
        sys.exit(1)
    work = args[0]
    only = None
    matroot = os.path.join(tempfile.gettempdir(), "vtr-mat")
    transcripts = {}
    rasterize = "--rasterize" in sys.argv[1:]
    for a in sys.argv[1:]:
        if a.startswith("--only="):
            only = set(a.split("=", 1)[1].split(","))
        elif a.startswith("--matroot="):
            matroot = a.split("=", 1)[1]
        elif a.startswith("--transcript="):
            k, v = a.split("=", 1)[1].split("=", 1)
            transcripts[k] = v
    os.makedirs(matroot, exist_ok=True)

    with open(os.path.join(work, "review_plan.json"), encoding="utf-8") as f:
        plan = json.load(f)

    out_path = os.path.join(work, "materials.json")
    materials = {}
    if only and os.path.exists(out_path):
        # Load-and-merge: a filtered re-run must never drop skipped folders.
        with open(out_path, encoding="utf-8") as f:
            materials = json.load(f)

    plan_fids = {e["folder_id"] for e in plan["folders"]}
    for k in transcripts:
        if k not in plan_fids:
            print(f"AVISO: --transcript={k}=... no corresponde a ninguna "
                  "carpeta del plan — la transcripción NO se adjuntará "
                  "(¿folder_id mal escrito?)", file=sys.stderr)

    for e in plan["folders"]:
        fid = e["folder_id"]
        if only and fid not in only:
            continue
        items = []
        srcs = ([e["deck"]] + e.get("deck_source_of", []) + e.get("evidence", [])
                if e["status"] == "review" else e.get("no_deck_files", []))
        for fr in srcs:
            process_file(fr, fid, items, matroot)
        if rasterize:
            for it in items:
                if it["kind"] == "pdf" and it.get("path"):
                    imgdir = os.path.splitext(it["path"])[0] + "_pages"
                    imgs, trunc, err = rasterize_pdf(it["path"], imgdir)
                    if imgs:
                        it["page_images"] = imgs
                        if trunc:
                            it["truncated_from"] = trunc
                    else:
                        it["note"] = ((it.get("note", "") + " | ") if it.get("note")
                                      else "") + f"rasterización falló: {err}"
        for it in items:
            if it["kind"] == "video" and fid in transcripts:
                it["transcript_path"] = transcripts[fid]
        materials[fid] = {"folder_name": e["folder_name"],
                          "student_name": e["student_name"],
                          "status": e["status"], "items": items}
        kinds = {}
        for it in items:
            kinds[it["kind"]] = kinds.get(it["kind"], 0) + 1
        print(f"{e['student_name'][:38]:38s} {kinds}", flush=True)

    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(materials, f, ensure_ascii=False, indent=2)
    errs = [(m["student_name"], it["label"], it["note"])
            for m in materials.values() for it in m["items"]
            if it["kind"] == "error"]
    print(f"\nwrote {out_path}: {len(materials)} carpetas, {len(errs)} ítems con error")
    for s, l, n in errs:
        print(f"  ERROR {s} | {l} | {n}", file=sys.stderr)
    vids = [(fid, it["label"]) for fid, m in materials.items()
            for it in m["items"]
            if it["kind"] == "video" and not it.get("transcript_path")]
    if vids:
        print(f"\n{len(vids)} video(s) sin transcripción — transcríbelos y "
              "vuelve a correr con --transcript=<folder_id>=<archivo.txt>:")
        for fid, label in vids:
            print(f"  {fid}  {label}")


if __name__ == "__main__":
    main()
