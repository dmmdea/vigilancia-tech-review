#!/usr/bin/env python3
"""Build the three-sheet results Excel (Ranking, Detalle, Meta) from the orchestrator's results JSON.

Usage:
    python make_excel.py <results.json> <output.xlsx> [--listing=<listing.json>]...

The final grade is computed HERE (single source of truth):
    final = 0.50*poc + 0.25*impacto + 0.25*comunicacion   (rounded to 0.01)
    disqualified -> final = 1.0
A reviewed row with missing/invalid/out-of-range scores is DOWNGRADED to
"no_revisado" with a visible reason — it never renders as a normal graded row
and never aborts the workbook (fairness: every submission stays visible).

Row identity is the Drive file id (`id` field) when present, so duplicate
filenames cannot corrupt the top-5. Ranking order: reviewed & not DQ (by final
desc) -> DQ -> not reviewed. Top 5 non-DQ rows are starred and highlighted.

--listing (repeatable): the drive_list.py JSON(s) for the folder and any
subfolders. When given, EVERY listed entry needs a results row (reviewed or
NO REVISADO) — except folders that were themselves passed as a --listing.
Any entry without a row fails the build (exit 2) naming the absent files —
the mechanical backstop for "nothing is silently dropped".
Exit codes: 0 ok · 1 usage (bad args, unknown option, unreadable results
JSON) · 2 empty results / duplicate id / unreadable or invalid listing /
completeness violation.
"""
import json
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Windows defaults std streams to cp1252; names and messages are non-ASCII.
for _s in (sys.stdout, sys.stderr):
    # Only touch the process's own console streams — never a stream an
    # importing caller substituted (reconfiguring theirs corrupts their file).
    if _s in (sys.__stdout__, sys.__stderr__) and hasattr(_s, "reconfigure"):
        _s.reconfigure(encoding="utf-8")

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TOP5_FILL = PatternFill("solid", fgColor="FDE68A")
DQ_FILL = PatternFill("solid", fgColor="FECACA")
NOREV_FILL = PatternFill("solid", fgColor="E5E7EB")
ANEXO_FILL = PatternFill("solid", fgColor="D1FAE5")    # leído en la revisión
REEMP_FILL = PatternFill("solid", fgColor="EDE9FE")    # versión reemplazada
THIN = Border(*[Side(style="thin", color="D1D5DB")] * 4)

RANK_COLS = [
    ("Posición", 9), ("Top 5", 7), ("Archivo", 38), ("Estudiante", 24),
    ("Herramienta", 26), ("Fecha lanz. declarada", 15),
    ("Fecha lanz. verificada", 15), ("Fuente verificación", 40),
    ("Confianza", 10), ("Edad (meses)", 9), ("¿Descalificado?", 13),
    ("Razón DQ", 30), ("PoC (50%)", 10), ("Impacto (25%)", 10),
    ("Comunicación (25%)", 12), ("Nota final", 10),
    ("Indicio IA (1-5)", 10),
    ("Flags revisión humana", 28),
    ("Estado", 16), ("Detalle estado", 30),
]

DETAIL_COLS = [
    ("Archivo", 38), ("Herramienta", 24), ("Slides leídas / total", 12),
    ("Justificación PoC", 60), ("Justificación Impacto", 60),
    ("Justificación Comunicación", 60),
    ("Feedback sugerido (borrador interno)", 60),
    ("Evidencia indicio IA", 50),
    ("Notas de evidencia", 60),
]


def normalize(r: dict) -> None:
    """Validate one result row in place: set r['_final'] or downgrade to no_revisado."""
    # flags is the one field this module appends to and joins — coerce any
    # LLM-drifted shape (null, bare string, list of dicts) to a list of strings.
    f = r.get("flags")
    r["flags"] = ([str(x) for x in f] if isinstance(f, list)
                  else [str(f)] if f else [])
    # Mechanical backstops for the launch-date rules (prompt-only otherwise):
    age = r.get("age_months")
    conf = r.get("verification_confidence")
    if isinstance(age, (int, float)):
        if 3.5 <= age <= 4.5 and "VERIFICAR FECHA" not in r["flags"]:
            r["flags"].append("VERIFICAR FECHA")
    elif conf in ("alta", "media") and r.get("status") == "revisado":
        # confident verification must yield a numeric age, else the DQ filter
        # silently never fires
        if "REVISAR MANUALMENTE" not in r["flags"]:
            r["flags"].append("REVISAR MANUALMENTE")
        if "EDAD SIN CALCULAR" not in r["flags"]:
            r["flags"].append("EDAD SIN CALCULAR")
    status = r.get("status")
    if status not in ("revisado", "no_revisado", "revisado_anexo",
                      "reemplazada"):
        r["status"] = "no_revisado"
        r.setdefault("status_reason", "estado ausente o inválido en results.json")
        r["_final"] = None
        return
    if status in ("no_revisado", "revisado_anexo", "reemplazada"):
        r["_final"] = None
        return
    if r.get("disqualified"):
        r["_final"] = 1.0
        return
    s = r.get("scores") or {}
    try:
        poc = float(s["poc"])
        imp = float(s["impacto"])
        com = float(s["comunicacion"])
    except (KeyError, TypeError, ValueError):
        r["status"] = "no_revisado"
        r["status_reason"] = "puntajes ausentes o inválidos en la revisión"
        r["flags"].append("REVISAR MANUALMENTE")
        r["_final"] = None
        return
    if any(not 1.0 <= v <= 5.0 for v in (poc, imp, com)):
        r["status"] = "no_revisado"
        r["status_reason"] = (f"puntaje fuera de rango 1.0-5.0 "
                              f"(poc={poc}, impacto={imp}, comunicacion={com})")
        r["flags"].append("REVISAR MANUALMENTE")
        r["_final"] = None
        return
    r["_final"] = round(0.50 * poc + 0.25 * imp + 0.25 * com, 2)


def sort_key(r: dict):
    status = r.get("status")
    dq = bool(r.get("disqualified"))
    final = r.get("_final")
    if status == "revisado":
        group = 1 if dq else 0
    else:
        group = {"revisado_anexo": 2, "reemplazada": 3}.get(status, 4)
    return (group, -(final if final is not None else 0.0))


def check_completeness(results: list, listing_paths: list) -> None:
    """EVERY listed entry must be accounted for: non-folder entries need a
    results row (reviewed or NO REVISADO); folder entries need either their own
    --listing (they were explored) or a results row (declared unexplored)."""
    expected = {}
    explored_folders = set()
    entries = []
    for p in listing_paths:
        try:
            with open(p, encoding="utf-8") as f:
                listing = json.load(f)
        except (OSError, json.JSONDecodeError) as e:
            print(f"ERROR: cannot read listing {p}: {e}", file=sys.stderr)
            sys.exit(2)
        if (not isinstance(listing, dict)
                or not isinstance(listing.get("entries"), list)
                or not listing.get("folder_id")
                or not all(isinstance(e, dict) for e in listing["entries"])):
            print(f"ERROR: {p} is not a drive_list.py listing "
                  "(needs folder_id and a list of entry objects).",
                  file=sys.stderr)
            sys.exit(2)
        explored_folders.add(listing["folder_id"])
        entries.extend(listing["entries"])
    for e in entries:
        if e.get("kind") == "folder" and e.get("id") in explored_folders:
            continue
        expected[e.get("id")] = e.get("name", "(sin nombre)")
    have = {r.get("id") for r in results if r.get("id")}
    missing = {i: n for i, n in expected.items() if i not in have}
    if missing:
        print("ERROR: results.json is missing rows for these listed entries "
              "(nothing may be silently dropped — every entry gets a row, "
              "reviewed or NO REVISADO):", file=sys.stderr)
        for i, n in missing.items():
            print(f"  - {n} (id {i})", file=sys.stderr)
        sys.exit(2)


def style_header(ws, cols):
    for i, (name, width) in enumerate(cols, 1):
        c = ws.cell(row=1, column=i, value=name)
        c.fill, c.font = HEADER_FILL, HEADER_FONT
        c.alignment = Alignment(wrap_text=True, vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"


def main() -> None:
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    listings = [a.split("=", 1)[1] for a in sys.argv[1:]
                if a.startswith("--listing=")]
    unknown = [a for a in sys.argv[1:]
               if a.startswith("--") and not a.startswith("--listing=")]
    if unknown:
        # A typo'd flag must never silently disarm the completeness gate.
        print(f"ERROR: unknown option(s): {' '.join(unknown)}", file=sys.stderr)
        sys.exit(1)
    if len(args) != 2:
        print(__doc__, file=sys.stderr)
        sys.exit(1)
    try:
        with open(args[0], encoding="utf-8") as f:
            data = json.load(f)
    except (OSError, json.JSONDecodeError) as e:
        print(f"ERROR: cannot read results JSON {args[0]}: {e}", file=sys.stderr)
        sys.exit(1)

    results = data.get("results", [])
    if not results:
        print("ERROR: results.json contains no results.", file=sys.stderr)
        sys.exit(2)
    if listings:
        check_completeness(results, listings)

    seen_ids = {}
    for idx, r in enumerate(results):
        rid = r.get("id")
        if rid:
            if rid in seen_ids:
                print(f"ERROR: duplicate id '{rid}' in results.json (rows "
                      f"{seen_ids[rid]} and {idx}) — one Drive file must have "
                      "exactly one row.", file=sys.stderr)
                sys.exit(2)
            seen_ids[rid] = idx
        r["_uid"] = rid or f"__row{idx}"
        normalize(r)
    results.sort(key=sort_key)

    rankable = [r for r in results
                if r.get("status") == "revisado"
                and not r.get("disqualified") and r["_final"] is not None]
    top5_uids = {r["_uid"] for r in rankable[:5]}
    rankable_uids = {r["_uid"] for r in rankable}
    # Flag a tie only when it actually straddles the top-5 cut (position 5 vs 6);
    # a tie wholly inside the top 5 contests nothing.
    if len(rankable) > 5 and rankable[5]["_final"] == rankable[4]["_final"]:
        cut = rankable[4]["_final"]
        for r in rankable:
            if r["_final"] == cut:
                r["flags"].append("EMPATE TOP5")

    wb = Workbook()
    ws = wb.active
    ws.title = "Ranking"
    style_header(ws, RANK_COLS)

    pos = 0
    for i, r in enumerate(results, start=2):
        reviewed = r.get("status") == "revisado"
        dq = bool(r.get("disqualified"))
        in_top5 = r["_uid"] in top5_uids
        if r["_uid"] in rankable_uids:
            pos += 1
            position = pos
        else:
            position = ""
        s = r.get("scores") or {}
        status = r.get("status")
        estado = {"revisado": "REVISADO",
                  "revisado_anexo": "REVISADO (ANEXO)",
                  "reemplazada": "REEMPLAZADA"}.get(status, "NO REVISADO")
        row = [
            position,
            "⭐ TOP 5" if in_top5 else "",
            r.get("file", ""), r.get("student", ""), r.get("tool", ""),
            r.get("declared_launch_date", ""), r.get("verified_launch_date", ""),
            r.get("verification_source", ""), r.get("verification_confidence", ""),
            r.get("age_months", ""),
            ("SÍ" if dq else "NO") if reviewed else "",
            r.get("dq_reason", ""),
            s.get("poc", ""), s.get("impacto", ""), s.get("comunicacion", ""),
            r["_final"] if r["_final"] is not None else "",
            r.get("indicio_ia", ""),
            ", ".join(r.get("flags", [])),
            estado,
            r.get("status_reason", ""),
        ]
        fill = (TOP5_FILL if in_top5
                else DQ_FILL if (reviewed and dq)
                else ANEXO_FILL if status == "revisado_anexo"
                else REEMP_FILL if status == "reemplazada"
                else NOREV_FILL if not reviewed else None)
        for j, v in enumerate(row, 1):
            c = ws.cell(row=i, column=j, value=v)
            c.border = THIN
            if fill:
                c.fill = fill
            if j in (13, 14, 15, 16):
                c.number_format = "0.00"

    ws2 = wb.create_sheet("Detalle")
    style_header(ws2, DETAIL_COLS)
    for i, r in enumerate(results, start=2):
        j = r.get("justification") or {}
        row = [
            r.get("file", ""), r.get("tool", ""),
            f'{r.get("pages_read", "?")} / {r.get("pages_total", "?")}',
            j.get("poc", ""), j.get("impacto", ""), j.get("comunicacion", ""),
            r.get("feedback_sugerido", ""),
            r.get("indicio_ia_evidencia", ""),
            r.get("evidence_notes", ""),
        ]
        for k, v in enumerate(row, 1):
            c = ws2.cell(row=i, column=k, value=v)
            c.border = THIN
            c.alignment = Alignment(wrap_text=True, vertical="top")

    meta = wb.create_sheet("Meta")
    meta["A1"], meta["B1"] = "Fecha de corrida", data.get("run_date", "")
    meta["A2"], meta["B2"] = "Carpeta Drive", data.get("folder_url", "")
    meta["A3"], meta["B3"] = "Regla de corte", ("> 4 meses desde lanzamiento verificado → 1.0; "
                                                "zona 3.5–4.5 meses lleva flag VERIFICAR FECHA")
    meta["A4"], meta["B4"] = "Ponderación", "PoC 50% · Impacto 25% · Comunicación 25%"
    meta["A6"], meta["B6"] = "Indicio IA (1-5)", ("señal ADVISORY de uso de IA sin filtro sobre el material "
                                                  "entregado (1=curado a mano, 5=volcado sin filtrar); NUNCA "
                                                  "es componente de la nota")
    meta["A7"], meta["B7"] = "Estados", ("REVISADO=fila calificada · REVISADO (ANEXO)=archivo leído dentro "
                                          "de la revisión integral del estudiante · REEMPLAZADA=versión "
                                          "anterior superada por una entrega más reciente · NO REVISADO=requiere humano")
    meta["A5"], meta["B5"] = "Generado por", ("vigilancia-tech-review (revisores de IA con "
                                              "contexto limpio; revisión humana "
                                              "requerida para nota oficial)")
    meta.column_dimensions["A"].width = 20
    meta.column_dimensions["B"].width = 80

    wb.save(args[1])
    print(f"OK {args[1]} ({len(results)} filas, top5={len(top5_uids)})")


if __name__ == "__main__":
    main()
