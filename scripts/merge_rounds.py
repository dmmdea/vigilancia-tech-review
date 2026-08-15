#!/usr/bin/env python3
"""Maestro multi-ronda: acumula las rondas de entregas en UN workbook sin
borrar jamás el histórico de rondas anteriores.

    python merge_rounds.py <maestro.xlsx> <ronda.xlsx> --round="Semana 2"

- Incorpora el Excel de la ronda (Ranking/Detalle/Meta de make_excel.py) al
  maestro como un trío de hojas propio de la ronda.
- Los títulos de hoja son DETERMINISTAS y a prueba de colisiones: Excel corta
  los títulos a 31 caracteres, así que dos rondas con nombres largos
  parecidos colapsarían en la misma hoja y una BORRARÍA a la otra (defecto
  real encontrado por revisión). Los nombres largos llevan un sufijo hash del
  nombre completo, y la hoja oculta "_Rondas" registra título↔nombre completo
  + orden de llegada. Si un título calculado ya pertenece a OTRA ronda según
  el registro, el script se niega (exit 2) en lugar de borrar historia.
- Re-entregar la MISMA ronda reemplaza SOLO su trío (refresh); las hojas de
  cualquier otra ronda no se tocan nunca.
- Hoja "Histórico": nota final por estudiante por ronda, en orden de llegada
  de las rondas. Se agrupa por la columna estable **Clave** (clave Canvas /
  id de carpeta) — el nombre visible del estudiante varía entre revisores
  (acentos, abreviaciones, códigos) y partiría el histórico; la clave no.
  El nombre mostrado es el de la ronda más reciente donde aparece.

Exit codes: 0 ok · 1 uso (args/round inválido) · 2 entrada ilegible, columnas
faltantes, o conflicto de títulos que borraría otra ronda.

MAX_PATH: escribe a un temporal corto y copia al destino con prefijo largo.
"""
import hashlib
import os
import re
import shutil
import sys
import tempfile
from copy import copy

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

for _s in (sys.stdout, sys.stderr):
    if _s in (sys.__stdout__, sys.__stderr__) and hasattr(_s, "reconfigure"):
        _s.reconfigure(encoding="utf-8")

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(bold=True, color="FFFFFF")
KINDS = ("Ranking", "Detalle", "Meta")
REGISTRY = "_Rondas"
ILLEGAL = re.compile(r"[\\/?*\[\]:]")


def longpath(p):
    if os.name != "nt":
        return p
    ap = os.path.abspath(p)
    if ap.startswith("\\\\?\\"):
        return ap
    if ap.startswith("\\\\"):
        return "\\\\?\\UNC\\" + ap[2:]
    return "\\\\?\\" + ap


def sheet_title(kind, rnd):
    """Deterministic, <=31 chars, collision-proof across DIFFERENT rounds:
    long round names get a 4-hex hash of the FULL name, so truncation can
    never make two rounds share a title."""
    base = f"{kind} - {rnd}"
    if len(base) <= 31:
        return base
    h = hashlib.sha256(rnd.encode("utf-8")).hexdigest()[:4]
    keep = max(31 - len(kind) - 3 - 5, 1)  # "kind - " + "~hash"; never <=0
    return f"{kind} - {rnd[:keep]}~{h}"


def load_registry(wb):
    """[(round_full_name, {kind: title})] in arrival order.

    A master written before the registry existed has no _Rondas sheet but
    DOES have per-round sheets — backfill from the 'Ranking - <ronda>'
    titles so the first post-upgrade merge doesn't rebuild a Histórico
    that silently drops every prior round (convergence finding 2)."""
    if REGISTRY not in wb.sheetnames:
        out = []
        for title in wb.sheetnames:
            if not title.startswith("Ranking - "):
                continue
            # only a REAL round sheet backfills — a TA's hand-added
            # "Ranking - consolidado" must not become a phantom round
            # (round-2 finding 2)
            hdr = [c.value for c in wb[title][1]]
            if not {"Estudiante", "Nota final", "Estado"} <= set(hdr):
                print(f"AVISO: la hoja '{title}' parece de ronda pero no "
                      "tiene las columnas de make_excel — se ignora en la "
                      "reconstrucción del registro.", file=sys.stderr)
                continue
            rnd = title[len("Ranking - "):]
            # the pre-registry version TRUNCATED titles to 31 chars, and
            # each kind prefix has a different length — probe by prefix,
            # not by exact recomposition (round-2 finding 1, Meta orphan)
            titles = {}
            for k in KINDS:
                want = f"{k} - {rnd}"
                # exact match wins OUTRIGHT; a prefix hit only counts when
                # UNIQUE — a disjunctive next() bound whichever round came
                # first in tab order and could hand this entry ANOTHER
                # round's sheets (round-3 finding 1: silent grade loss)
                if want in wb.sheetnames:
                    titles[k] = want
                    continue
                cands = [t for t in wb.sheetnames if t.startswith(want)]
                if len(cands) == 1:
                    titles[k] = cands[0]
                elif cands:
                    print(f"AVISO: varias hojas coinciden con '{want}' "
                          f"({cands}) — ambigua, se omite del registro "
                          "reconstruido.", file=sys.stderr)
            out.append((rnd, titles))
        if out:
            print(f"AVISO: maestro sin hoja de registro '{REGISTRY}' — "
                  f"{len(out)} ronda(s) reconstruida(s) desde los títulos "
                  "de hoja existentes (posiblemente truncados a 31 chars "
                  "por la versión anterior).", file=sys.stderr)
        return out
    out = []
    for row in wb[REGISTRY].iter_rows(min_row=2, values_only=True):
        if row and row[0]:
            out.append((row[0], {k: t for k, t in zip(KINDS, row[1:4]) if t}))
    return out


def save_registry(wb, entries):
    if REGISTRY in wb.sheetnames:
        del wb[REGISTRY]
    ws = wb.create_sheet(REGISTRY)
    ws.sheet_state = "hidden"
    ws.append(["Ronda", *KINDS])
    for name, titles in entries:
        ws.append([name] + [titles.get(k, "") for k in KINDS])


def copy_sheet(src_ws, dst_wb, title):
    ws = dst_wb.create_sheet(title=title)
    for row in src_ws.iter_rows():
        for cell in row:
            c = ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                c.font = copy(cell.font)
                c.fill = copy(cell.fill)
                c.border = copy(cell.border)
                c.alignment = copy(cell.alignment)
                c.number_format = cell.number_format
    for col, dim in src_ws.column_dimensions.items():
        if dim.width:
            ws.column_dimensions[col].width = dim.width
    if src_ws.freeze_panes:
        ws.freeze_panes = src_ws.freeze_panes
    return ws


def rebuild_historico(wb, registry):
    """Grade per student per round, arrival-ordered, keyed on the stable
    'Clave' column (falls back to the visible name for pre-Clave rounds)."""
    if "Histórico" in wb.sheetnames:
        del wb["Histórico"]
    rounds = [name for name, _t in registry]
    data = {}          # key -> {"name": display, "grades": {round: final}}
    order = []
    keyed, name_keyed = [], []
    for name, titles in registry:
        t = titles.get("Ranking")
        if not t or t not in wb.sheetnames:
            # a REGISTERED round whose sheet is gone is registry/workbook
            # divergence, not "no data" — its Histórico column will be
            # blank; say so instead of silently succeeding (finding 4)
            print(f"AVISO: la ronda registrada '{name}' no tiene su hoja "
                  f"'{t or '(sin título de Ranking registrado)'}' en el maestro (¿renombrada o borrada a mano?) — "
                  "su columna del Histórico quedará vacía.", file=sys.stderr)
            continue
        ws = wb[t]
        hdr = [c.value for c in ws[1]]
        try:
            i_est = hdr.index("Estudiante")
            i_fin = hdr.index("Nota final")
            i_st = hdr.index("Estado")
        except ValueError:
            print(f"AVISO: la hoja '{t}' de la ronda '{name}' no tiene las "
                  "columnas esperadas (Estudiante/Nota final/Estado) — su "
                  "columna del Histórico quedará vacía.", file=sys.stderr)
            continue
        i_key = hdr.index("Clave") if "Clave" in hdr else None
        (keyed if i_key is not None else name_keyed).append(name)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[i_st] != "REVISADO":
                continue
            est = (row[i_est] or "").strip()
            key = ((row[i_key] or "").strip() if i_key is not None else "") or est
            if not key:
                continue
            if key in data and name in data[key]["grades"]:
                # two REVISADO rows in the SAME round collapsing onto one
                # key (blank Clave + shared display name) would silently
                # drop a grade — disambiguate with a TUPLE key, which can
                # never equal a real string Clave or display name (round-3
                # finding 3: an "X (2)" string key merged the duplicate
                # into a DIFFERENT real student), and label the extra row
                # so the two aren't visually identical.
                print(f"AVISO: dos filas REVISADO de la ronda '{name}' "
                      f"comparten la clave '{key}' — se separan en filas "
                      "distintas del Histórico; verificar la identidad de "
                      "esos estudiantes.", file=sys.stderr)
                base, n = key, 2
                while (base, n) in data and name in data[(base, n)]["grades"]:
                    n += 1
                key = (base, n)
                if key not in data:
                    data[key] = {"name": f"{est} (fila duplicada {n})",
                                 "grades": {}}
                    order.append(key)
                data[key]["grades"][name] = row[i_fin]
                continue    # keep the duplicate's marked display name
            if key not in data:
                data[key] = {"name": est, "grades": {}}
                order.append(key)
            data[key]["grades"][name] = row[i_fin]
            if est:
                data[key]["name"] = est     # latest round's display name wins
    if keyed and name_keyed:
        # mixed keying splits every student whose Clave-keyed and name-keyed
        # rows don't coincide — visible only here, so warn here (finding 5)
        print("AVISO: el Histórico mezcla rondas CON columna 'Clave' "
              f"({', '.join(keyed)}) y SIN ella ({', '.join(name_keyed)}) — "
              "un mismo estudiante puede aparecer partido en dos filas. "
              "Regenera las rondas viejas con el make_excel actual para "
              "unificar.", file=sys.stderr)
    ws = wb.create_sheet("Histórico", 0)
    hdr = ["Estudiante"] + [f"Nota {r}" for r in rounds]
    for j, h in enumerate(hdr, 1):
        c = ws.cell(row=1, column=j, value=h)
        c.fill, c.font = HEADER_FILL, HEADER_FONT
        c.alignment = Alignment(wrap_text=True, vertical="center")
    ws.column_dimensions["A"].width = 30
    for j in range(2, len(hdr) + 1):
        ws.column_dimensions[get_column_letter(j)].width = 14
    ordered = sorted(order, key=lambda k: str.casefold(data[k]["name"]))
    for i, key in enumerate(ordered, 2):
        ws.cell(row=i, column=1, value=data[key]["name"])
        for j, rname in enumerate(rounds, 2):
            v = data[key]["grades"].get(rname)
            if v is not None:
                c = ws.cell(row=i, column=j, value=v)
                c.number_format = "0.00"
    ws.freeze_panes = "B2"
    return len(order), rounds


def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    rnd = None
    for a in sys.argv[1:]:
        if a.startswith("--round="):
            rnd = a.split("=", 1)[1].strip()
    if len(args) != 2 or not rnd:
        print(__doc__, file=sys.stderr)
        sys.exit(1)
    if ILLEGAL.search(rnd):
        print(f"ERROR: el nombre de ronda '{rnd}' contiene caracteres no "
              "permitidos en títulos de hoja de Excel (\\ / ? * [ ] :) — "
              "usa otro nombre.", file=sys.stderr)
        sys.exit(1)
    if rnd.endswith("'"):
        # Excel forbids sheet titles ENDING with an apostrophe (the title
        # always starts with the kind prefix, so a leading one is fine)
        print(f"ERROR: el nombre de ronda '{rnd}' no puede terminar con "
              "apóstrofo (regla de Excel para títulos de hoja).",
              file=sys.stderr)
        sys.exit(1)
    master_path, round_path = args

    try:
        src = load_workbook(longpath(round_path))
    except Exception as e:
        print(f"ERROR: no se pudo leer el Excel de la ronda: {e}",
              file=sys.stderr)
        sys.exit(2)
    for needed in KINDS:
        if needed not in src.sheetnames:
            print(f"ERROR: al Excel de la ronda le falta la hoja '{needed}'.",
                  file=sys.stderr)
            sys.exit(2)
    hdr = [c.value for c in src["Ranking"][1]]
    for col in ("Estudiante", "Nota final", "Estado"):
        if col not in hdr:
            print(f"ERROR: la hoja Ranking de la ronda no tiene la columna "
                  f"'{col}' — ¿se generó con make_excel.py?", file=sys.stderr)
            sys.exit(2)
    if "Clave" not in hdr:
        print("AVISO: la ronda no trae columna 'Clave' — el Histórico usará "
              "el nombre visible del estudiante, que puede variar entre "
              "rondas.", file=sys.stderr)

    if os.path.exists(longpath(master_path)):
        wb = load_workbook(longpath(master_path))
    else:
        wb = Workbook()
        wb.remove(wb.active)

    registry = load_registry(wb)

    # Resolve round IDENTITY against reconstructed (pre-registry) names,
    # which are 31-char TRUNCATED remnants of the real round name
    # (round-2 findings 1 and 3):
    #  * hash-form titles round-trip exactly (the hash pins the full name)
    #    → same round: ADOPT rnd as the entry's name so refresh works;
    #  * a truncated-remnant prefix match is AMBIGUOUS (same round being
    #    refreshed, or a genuinely different round sharing the prefix) —
    #    proceeding would either duplicate the week or clobber another:
    #    refuse with both exits spelled out.
    if not any(name == rnd for name, _t in registry):
        resolved = []
        adopted = False
        adopt_old_titles = []
        for name, btitles in registry:
            rk = btitles.get("Ranking", "")
            if (not adopted and name != rnd and rk
                    and rk == sheet_title("Ranking", rnd)):
                # EXACT title match only: the reconstructed hash-form name
                # round-trips byte-identical. Casefolding here would make a
                # case-variant round name silently adopt-and-refresh instead
                # of refusing (the F3 guarantee). At most ONE entry may
                # adopt (round-3 finding 2: a duplicate adopt erased a
                # round and doubled the Histórico column).
                adopted = True
                adopt_old_titles = list(btitles.values())
                name = rnd
                print(f"AVISO: la ronda reconstruida '{rk[10:]}' es la "
                      f"misma que '{rnd}' (título idéntico) — se refresca.",
                      file=sys.stderr)
            elif (name != rnd and len(rk) == 31 and "~" not in rk
                    and rnd.startswith(name)):
                # Before refusing, use the Meta title to decide what the
                # Ranking truncation hid (Meta keeps 3 more name chars):
                # a complete Meta name, or an extended prefix rnd does NOT
                # match, PROVES this is a different round (round-3
                # finding 4: a legitimate new round was refused, and the
                # guidance would have overwritten the old round's grades).
                mt = btitles.get("Meta", "")
                mname = mt[len("Meta - "):] if mt else ""
                if mt and len(mt) < 31 and mname == rnd and not adopted:
                    # complete Meta name EQUALS rnd → PROVES the same round
                    # (round-4 finding: assuming "complete → different"
                    # dead-ended a 22/23-char round's own refresh and the
                    # clash message's advice then duplicated the week)
                    adopted = True
                    adopt_old_titles = list(btitles.values())
                    name = rnd
                    print(f"AVISO: la ronda reconstruida '{rk[10:]}' es la "
                          f"misma que '{rnd}' (nombre completo en la hoja "
                          "Meta) — se refresca.", file=sys.stderr)
                elif mt and len(mt) < 31:
                    pass          # complete name != rnd → different round
                elif mname and not rnd.startswith(mname):
                    pass          # extended prefix disproves identity
                else:
                    print(f"ERROR: el maestro contiene una ronda "
                          f"reconstruida '{name}' (título truncado por la "
                          "versión anterior) que coincide con el inicio de "
                          f"'{rnd}'. No es posible saber mecánicamente si "
                          "es la misma ronda. Si ES la misma, re-mergéala "
                          f"con --round=\"{name}\"; si es una ronda "
                          "DISTINTA, usa un nombre que no empiece igual, o "
                          "regenera el maestro desde cero con los Excel "
                          "por ronda. NO se modificó nada.", file=sys.stderr)
                    sys.exit(2)
            resolved.append((name, btitles))
        registry = resolved
        # the adopted entry's OLD sheets (plain-truncated titles) must be
        # replaced, not orphaned, when the new titles differ (Meta-adopt)
        for old_t in adopt_old_titles:
            if old_t and old_t.casefold() not in                     {t.casefold() for t in
                     (sheet_title(k, rnd) for k in KINDS)}:
                if old_t in wb.sheetnames:
                    del wb[old_t]
    if sum(1 for name, _t in registry if name == rnd) > 1:
        print(f"ERROR: el registro contiene la ronda '{rnd}' más de una "
              "vez — registro corrupto; repara la hoja _Rondas antes de "
              "continuar. NO se modificó nada.", file=sys.stderr)
        sys.exit(2)

    titles = {k: sheet_title(k, rnd) for k in KINDS}

    # refuse to touch a title that the registry attributes to ANOTHER round —
    # deleting it would destroy that round's history. Excel sheet titles are
    # CASE-INSENSITIVE-unique (openpyxl silently renames on collision), so
    # every comparison here must casefold (convergence finding 3).
    ours_cf = {t.casefold() for t in titles.values()}
    for other_name, other_titles in registry:
        if other_name == rnd:
            continue
        clash = ours_cf & {t.casefold() for t in other_titles.values()}
        if clash:
            print(f"ERROR: el título de hoja {sorted(clash)} ya pertenece a "
                  f"la ronda '{other_name}' — nombres de ronda demasiado "
                  "parecidos; elige un nombre distinto. NO se borró nada.",
                  file=sys.stderr)
            sys.exit(2)

    replaced = any(name == rnd for name, _t in registry)
    for kind in KINDS:
        t = titles[kind]
        for existing in list(wb.sheetnames):
            if existing.casefold() == t.casefold():
                del wb[existing]
        ws = copy_sheet(src[kind], wb, t)
        if ws.title != t:
            # openpyxl renamed the sheet — the registry would point at a
            # title that doesn't exist and the round would silently vanish
            print(f"ERROR: openpyxl renombró la hoja '{t}' a '{ws.title}' "
                  "(colisión de títulos no detectada) — abortando sin "
                  "guardar.", file=sys.stderr)
            sys.exit(2)

    if replaced:
        registry = [(n, titles if n == rnd else t) for n, t in registry]
    else:
        registry.append((rnd, titles))
    save_registry(wb, registry)

    n_students, rounds = rebuild_historico(wb, registry)

    # never truncate the irreplaceable master in place: stage next to it,
    # then atomically swap (os.replace) so a crash mid-write leaves the
    # previous master intact
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.close()
    staged = longpath(master_path + ".tmp~")
    try:
        wb.save(tmp.name)
        shutil.copyfile(tmp.name, staged)
        os.replace(staged, longpath(master_path))
    except OSError as exc:
        # the staged copy holds student grades in the Drive-synced delivery
        # folder — never leave it behind; and a write failure is exit 2,
        # not the exit-1 "bad arguments" contract (round-2 finding 4)
        print(f"ERROR: no se pudo escribir el maestro: {exc}\n"
              "¿Está el archivo abierto en Excel? Ciérralo y reintenta. "
              "El maestro anterior quedó intacto.", file=sys.stderr)
        sys.exit(2)
    finally:
        for leftover in (tmp.name, staged):
            try:
                os.unlink(leftover)
            except FileNotFoundError:
                pass
            except OSError:
                if leftover == staged:
                    # the staged copy carries student grades in the
                    # Drive-synced folder — never leave it behind silently
                    print(f"AVISO: no se pudo borrar el temporal "
                          f"'{master_path}.tmp~' — bórralo a mano "
                          "(contiene notas de estudiantes).",
                          file=sys.stderr)
    print(f"OK maestro '{os.path.basename(master_path)}': ronda '{rnd}' "
          f"{'reemplazada' if replaced else 'agregada'} · rondas presentes: "
          f"{rounds} · estudiantes en Histórico: {n_students}")


if __name__ == "__main__":
    main()
